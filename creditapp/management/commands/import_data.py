from django.core.management import BaseCommand
from creditapp.models import Customer, Loan
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "This command imports customer and loan data from Excel files using background workers."

    def handle(self, *args, **options):
        # Import customer data in a background task
        import_customers()

        # Import loan data in a background task
        import_loans()

        self.stdout.write("Data import tasks started successfully.")


def import_customers():
    customer_workbook = load_workbook("data/customer_data.xlsx")
    customer_sheet = customer_workbook.active
    for row in customer_sheet.iter_rows(min_row=2):
        # Extract values and create Customer object
        try:
            if row[0].value == None:
                raise ValueError("None Value found.")
            customer = Customer.objects.create(
                customer_id=row[0].value,
                first_name=row[1].value,
                last_name=row[2].value,
                age=row[3].value,
                phone_number=row[4].value,
                monthly_salary=row[5].value,
                approved_limit=row[6].value
            )
            customer.save()
            print(f"Customer Id {row[0].value} Saved Successfully.")
        except BaseException as e:
            print(e)


def import_loans():
    loan_workbook = load_workbook("data/loan_data.xlsx")
    loan_sheet = loan_workbook.active
    for row in loan_sheet.iter_rows(min_row=2):
        try:
            customer = Customer.objects.get(
                customer_id=row[0].value
            )
            loan = Loan.objects.create(
                customer=customer,
                loan_id=row[1].value,
                loan_amount=row[2].value,
                tenure=row[3].value,
                interest_rate=row[4].value,
                monthly_repayment=row[5].value,
                emis_paid_on_time=row[6].value,
                start_date=row[7].value,
                end_date=row[8].value
            )
            loan.save()
            print(f"Loan Id {row[0].value} Saved Successfully.")
        except BaseException as e:
            print(e)
